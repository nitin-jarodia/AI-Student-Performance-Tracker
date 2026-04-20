# routes/bulk.py — Excel/CSV bulk upload for scores & students

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import Response

from app.dependencies.auth import CurrentUser, require_teacher
from app.services.audit import client_ip_from_request, log_action
from app.services.notification_service import dispatch_low_grade_alert_async
from app.services.rbac import ROLE_ADMIN, ROLE_TEACHER
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.fixed_subjects import FIXED_SUBJECTS_BY_ID, SUBJECT_HEADER_TO_ID
from app.models.models import Performance, Student, Subject

router = APIRouter(prefix="/bulk", tags=["Bulk Upload"])

MAX_PREVIEW_ROWS = 5


def _actor_role(user: CurrentUser) -> str:
    return ROLE_ADMIN if user.role == ROLE_ADMIN else ROLE_TEACHER


def norm_header(h: Any) -> str:
    return "_".join(str(h).strip().lower().split())


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [norm_header(c) for c in df.columns]
    return df


async def read_file_to_dataframe(file: UploadFile) -> pd.DataFrame:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    name = (file.filename or "").lower()
    bio = BytesIO(raw)
    try:
        if name.endswith(".csv"):
            return pd.read_csv(bio)
        return pd.read_excel(bio, engine="openpyxl")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}") from e


def parse_exam_date(val: Any) -> date:
    if pd.isna(val):
        raise ValueError("exam_date is empty")
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        raise ValueError("exam_date is empty")
    # Excel serial number
    if s.replace(".", "").isdigit():
        try:
            dt = pd.to_datetime(float(s), unit="d", origin="1899-12-30")
            return dt.date()
        except Exception:
            pass
    try:
        ts = pd.to_datetime(s, dayfirst=False)
        return ts.date()
    except Exception as e:
        raise ValueError(f"invalid exam_date: {val}") from e


def normalize_exam_type(raw: Any) -> str:
    if pd.isna(raw):
        raise ValueError("exam_type is empty")
    s = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "unittest": "unit_test",
        "unit_test": "unit_test",
        "mid_term": "midterm",
        "midterm": "midterm",
        "final_exam": "final",
        "final": "final",
        "quiz": "quiz",
        "assignment": "assignment",
        "practical": "practical",
    }
    return aliases.get(s, s)


def subject_columns_map(df: pd.DataFrame) -> Dict[int, str]:
    """subject_id -> first matching column name in dataframe."""
    found: Dict[int, str] = {}
    for col in df.columns:
        sid = SUBJECT_HEADER_TO_ID.get(col)
        if sid is not None and sid not in found:
            found[sid] = col
    return found


def parse_score_cell(val: Any) -> Optional[float]:
    if pd.isna(val):
        return None
    if isinstance(val, str) and not val.strip():
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(f"invalid score: {val}")


def scores_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    headers = [
        "roll_number",
        "Mathematics",
        "Science",
        "English",
        "Social Studies",
        "Computer Science",
        "exam_type",
        "exam_date",
    ]
    ws.append(headers)
    ws.append(["R001", 85, 78, 92, 88, 95, "unit_test", "2025-04-19"])

    ins = wb.create_sheet("Instructions")
    ins.append(["Column", "Description"])
    ins.append(["roll_number", "Student roll number; must exist in the system."])
    ins.append(["Mathematics … Computer Science", "Numeric score (0–100 typical). Leave blank to skip a subject."])
    ins.append(["exam_type", "One of: unit_test, midterm, final (or quiz, assignment, practical)."])
    ins.append(["exam_date", "Exam date (YYYY-MM-DD)."])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def students_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = [
        "name",
        "email",
        "roll_number",
        "class_name",
        "section",
        "parent_name",
        "parent_phone",
        "parent_email",
    ]
    ws.append(headers)
    ws.append(["Aarav Patel", "aarav@email.com", "R001", "10", "A", "Raj Patel", "9876543210", ""])

    ins = wb.create_sheet("Instructions")
    ins.append(["Column", "Description"])
    ins.append(["name", "Student full name (required)."])
    ins.append(["email", "Unique email (optional). Duplicates are skipped."])
    ins.append(["roll_number", "Unique roll number (required for new rows)."])
    ins.append(["class_name / section", "Class and section identifiers."])
    ins.append(["parent_*", "Guardian contact fields (optional)."])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@router.get("/template/scores")
def download_scores_template(_: CurrentUser = Depends(require_teacher)):
    data = scores_template_bytes()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="scores_template.xlsx"'},
    )


@router.get("/template/students")
def download_students_template(_: CurrentUser = Depends(require_teacher)):
    data = students_template_bytes()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="students_template.xlsx"'},
    )


@router.post("/preview-scores")
async def preview_scores(
    file: UploadFile = File(...),
    override_exam_type: Optional[str] = Form(None),
    override_exam_date: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: CurrentUser = Depends(require_teacher),
):
    df = normalize_columns(await read_file_to_dataframe(file))
    if "roll_number" not in df.columns:
        raise HTTPException(status_code=400, detail="Missing required column: roll_number")

    rolls = df["roll_number"].astype(str).str.strip()
    unique_rolls = rolls[rolls != ""].unique().tolist()

    by_roll = {
        s.roll_number: s.id
        for s in db.query(Student).filter(Student.roll_number.in_(unique_rolls)).all()
    }
    missing = [r for r in unique_rolls if r and r not in by_roll]

    preview_df = df.head(MAX_PREVIEW_ROWS)
    preview_rows = []
    for _, row in preview_df.iterrows():
        preview_rows.append({k: (None if pd.isna(v) else v) for k, v in row.items()})

    return {
        "total_rows": int(len(df)),
        "columns": list(df.columns),
        "preview_rows": preview_rows,
        "roll_numbers_found": len(unique_rolls) - len(missing),
        "roll_numbers_missing": missing,
        "override_exam_type": override_exam_type,
        "override_exam_date": override_exam_date,
    }


@router.post("/preview-students")
async def preview_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: CurrentUser = Depends(require_teacher),
):
    df = normalize_columns(await read_file_to_dataframe(file))
    required = {"name", "roll_number", "class_name", "section"}
    if not required.issubset(set(df.columns)):
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns; need all of: {sorted(required)}",
        )

    rolls_in_file: List[str] = []
    emails_in_file: List[str] = []
    for _, row in df.iterrows():
        r = str(row.get("roll_number", "")).strip()
        if r:
            rolls_in_file.append(r)
        e = row.get("email")
        if e is not None and not pd.isna(e) and str(e).strip():
            emails_in_file.append(str(e).strip().lower())

    existing_rolls = {
        r[0]
        for r in db.query(Student.roll_number).filter(Student.roll_number.in_(rolls_in_file)).all()
        if r[0]
    }
    existing_emails = set()
    if emails_in_file:
        existing_emails = {
            (e[0] or "").lower()
            for e in db.query(Student.email).filter(Student.email.in_(emails_in_file)).all()
            if e[0]
        }

    seen_roll: Dict[str, int] = {}
    seen_email: Dict[str, int] = {}
    dup_in_file_roll: List[str] = []
    dup_in_file_email: List[str] = []

    for idx, row in df.iterrows():
        r = str(row.get("roll_number", "")).strip()
        if r:
            if r in seen_roll:
                dup_in_file_roll.append(r)
            else:
                seen_roll[r] = int(idx)
        em = row.get("email")
        if em is not None and not pd.isna(em) and str(em).strip():
            el = str(em).strip().lower()
            if el in seen_email:
                dup_in_file_email.append(el)
            else:
                seen_email[el] = int(idx)

    preview_df = df.head(MAX_PREVIEW_ROWS)
    preview_rows = []
    for _, row in preview_df.iterrows():
        preview_rows.append({k: (None if pd.isna(v) else v) for k, v in row.items()})

    would_skip = 0
    for _, row in df.iterrows():
        r = str(row.get("roll_number", "")).strip()
        em = row.get("email")
        el = str(em).strip().lower() if em is not None and not pd.isna(em) and str(em).strip() else ""
        if r in existing_rolls or (el and el in existing_emails):
            would_skip += 1
            continue
        if r in dup_in_file_roll or (el and el in dup_in_file_email):
            would_skip += 1

    return {
        "total_rows": int(len(df)),
        "columns": list(df.columns),
        "preview_rows": preview_rows,
        "duplicate_rolls_in_file": sorted(set(dup_in_file_roll)),
        "duplicate_emails_in_file": sorted(set(dup_in_file_email)),
        "existing_roll_clash_count": sum(1 for r in rolls_in_file if r in existing_rolls),
        "existing_email_clash_count": sum(1 for e in emails_in_file if e in existing_emails),
        "rows_likely_skipped": would_skip,
    }


@router.post("/upload-scores")
async def upload_scores(
    request: Request,
    file: UploadFile = File(...),
    override_exam_type: Optional[str] = Form(None),
    override_exam_date: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_teacher),
):
    df = normalize_columns(await read_file_to_dataframe(file))
    if "roll_number" not in df.columns:
        raise HTTPException(status_code=400, detail="Missing required column: roll_number")

    sub_map = subject_columns_map(df)
    if not sub_map:
        raise HTTPException(
            status_code=400,
            detail="No subject columns found. Use math/mathematics, science, english, social_studies, computer, etc.",
        )

    use_exam_type = override_exam_type.strip() if override_exam_type and override_exam_type.strip() else None
    use_exam_date: Optional[date] = None
    if override_exam_date and str(override_exam_date).strip():
        try:
            use_exam_date = parse_exam_date(str(override_exam_date).strip())
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e

    if not use_exam_type and "exam_type" not in df.columns:
        raise HTTPException(
            status_code=400,
            detail="Provide override_exam_type or include exam_type in the file.",
        )
    if not use_exam_date and "exam_date" not in df.columns:
        raise HTTPException(
            status_code=400,
            detail="Provide override_exam_date or include exam_date in the file.",
        )

    # One query for all rolls in file (avoid N round-trips)
    rolls_unique = [
        str(x).strip()
        for x in df["roll_number"].tolist()
        if str(x).strip()
    ]
    rolls_unique = list(dict.fromkeys(rolls_unique))
    student_rows = db.query(Student).filter(Student.roll_number.in_(rolls_unique)).all()
    by_roll: Dict[str, Student] = {s.roll_number: s for s in student_rows}

    # When UI sends fixed exam + date, preload existing scores once (avoid per-cell SELECTs)
    db_existing_keys: Optional[Set[Tuple[int, int, Any, str]]] = None
    if use_exam_type and use_exam_date:
        et0 = normalize_exam_type(use_exam_type)
        ed0 = use_exam_date
        sid_list = [s.id for s in student_rows]
        if sid_list:
            rows = (
                db.query(Performance)
                .filter(
                    Performance.student_id.in_(sid_list),
                    Performance.exam_date == ed0,
                    Performance.exam_type == et0,
                )
                .all()
            )
            db_existing_keys = {(p.student_id, p.subject_id, p.exam_date, p.exam_type) for p in rows}

    errors: List[Dict[str, Any]] = []
    success_count = 0
    failed_rows = 0
    skipped_count = 0

    pending: List[Performance] = []
    batch_keys: Set[Tuple[int, int, Any, str]] = set()

    for idx, row in df.iterrows():
        row_label = int(idx) + 2 if isinstance(idx, int) else str(idx)
        roll = str(row.get("roll_number", "")).strip()
        if not roll:
            failed_rows += 1
            errors.append({"row": row_label, "detail": "empty roll_number"})
            continue

        stu = by_roll.get(roll)
        if not stu:
            failed_rows += 1
            errors.append({"row": row_label, "detail": f"roll_number not found: {roll}"})
            continue

        try:
            et = normalize_exam_type(use_exam_type) if use_exam_type else normalize_exam_type(row.get("exam_type"))
            ed = use_exam_date if use_exam_date else parse_exam_date(row.get("exam_date"))
        except ValueError as e:
            failed_rows += 1
            errors.append({"row": row_label, "detail": str(e)})
            continue

        scores_in_row = 0
        for sid, col in sub_map.items():
            try:
                sc = parse_score_cell(row.get(col))
            except ValueError as e:
                errors.append({"row": row_label, "detail": f"{col}: {e}"})
                continue
            if sc is None:
                continue
            scores_in_row += 1
            key = (stu.id, sid, ed, et)
            if key in batch_keys:
                skipped_count += 1
                continue
            if db_existing_keys is not None:
                if key in db_existing_keys:
                    skipped_count += 1
                    continue
            else:
                existing = (
                    db.query(Performance)
                    .filter(
                        Performance.student_id == stu.id,
                        Performance.subject_id == sid,
                        Performance.exam_date == ed,
                        Performance.exam_type == et,
                    )
                    .first()
                )
                if existing:
                    skipped_count += 1
                    continue
            batch_keys.add(key)
            pending.append(
                Performance(
                    student_id=stu.id,
                    subject_id=sid,
                    score=sc,
                    max_score=100.0,
                    exam_type=et,
                    exam_date=ed,
                )
            )
            success_count += 1

        if scores_in_row == 0:
            failed_rows += 1
            errors.append({"row": row_label, "detail": f"no scores entered for roll {roll}"})

    try:
        for p in pending:
            db.add(p)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}") from e

    log_action(
        user.email,
        _actor_role(user),
        "BULK_UPLOAD",
        target_type="performance",
        target_id=None,
        detail={"kind": "scores", "success_count": success_count, "filename": file.filename},
        ip_address=client_ip_from_request(request),
    )

    # Dispatch low-grade alerts for any rows that fell below 40% (dedup per student+subject).
    alerts_queued = 0
    if pending:
        subject_name_by_id: Dict[int, str] = {
            s.id: s.name for s in db.query(Subject).all()
        }
        student_by_id: Dict[int, Student] = {s.id: s for s in db.query(Student).all()}
        already: Set[Tuple[int, int]] = set()
        for p in pending:
            try:
                pct = (p.score / p.max_score) * 100 if p.max_score else 0.0
            except Exception:
                continue
            if pct >= 40.0:
                continue
            key = (p.student_id, p.subject_id)
            if key in already:
                continue
            already.add(key)
            student = student_by_id.get(p.student_id)
            if not student:
                continue
            dispatch_low_grade_alert_async(
                student,
                subject_name=subject_name_by_id.get(p.subject_id, "(subject)"),
                score=round(pct, 2),
                threshold_pct=40.0,
            )
            alerts_queued += 1

    return {
        "success_count": success_count,
        "failed_count": failed_rows,
        "skipped_count": skipped_count,
        "errors": errors[:200],
        "errors_truncated": len(errors) > 200,
        "alerts_queued": alerts_queued,
    }


@router.post("/upload-students")
async def upload_students(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: CurrentUser = Depends(require_teacher),
):
    df = normalize_columns(await read_file_to_dataframe(file))
    required = {"name", "roll_number", "class_name", "section"}
    if not required.issubset(set(df.columns)):
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns; need: {sorted(required)}",
        )

    inserted = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_label = int(idx) + 2 if isinstance(idx, int) else str(idx)
        name = str(row.get("name", "")).strip()
        roll = str(row.get("roll_number", "")).strip()
        cls = str(row.get("class_name", "")).strip()
        sec = str(row.get("section", "")).strip()

        if not name or not roll or not cls or not sec:
            skipped += 1
            errors.append({"row": row_label, "detail": "missing name, roll_number, class_name, or section"})
            continue

        em_raw = row.get("email")
        email = None
        if em_raw is not None and not pd.isna(em_raw) and str(em_raw).strip():
            email = str(em_raw).strip()

        if db.query(Student).filter(Student.roll_number == roll).first():
            skipped += 1
            errors.append({"row": row_label, "detail": f"duplicate roll_number: {roll}"})
            continue

        if email and db.query(Student).filter(Student.email == email).first():
            skipped += 1
            errors.append({"row": row_label, "detail": f"duplicate email: {email}"})
            continue

        pn = row.get("parent_name")
        parent_name = None if pn is None or pd.isna(pn) else str(pn).strip()
        pp = row.get("parent_phone")
        parent_phone = None if pp is None or pd.isna(pp) else str(pp).strip()
        pe = row.get("parent_email")
        parent_email = None if pe is None or pd.isna(pe) else str(pe).strip()

        db.add(
            Student(
                name=name,
                email=email,
                roll_number=roll,
                class_name=cls,
                section=sec,
                parent_name=parent_name or None,
                parent_phone=parent_phone or None,
                parent_email=parent_email or None,
            )
        )
        inserted += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}") from e

    log_action(
        user.email,
        _actor_role(user),
        "BULK_UPLOAD",
        target_type="student",
        target_id=None,
        detail={"kind": "students", "inserted": inserted, "filename": file.filename},
        ip_address=client_ip_from_request(request),
    )

    return {
        "inserted": inserted,
        "skipped": skipped,
        "errors": errors[:200],
        "errors_truncated": len(errors) > 200,
    }
